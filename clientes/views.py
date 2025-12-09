from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, Iterable, List
import unicodedata

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Sum
from django.http import HttpRequest, HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
@login_required
def acesso_negado(request: HttpRequest) -> HttpResponse:
    return render(request, "clientes/acesso_negado.html", status=403)

from openpyxl import Workbook, load_workbook

from .forms import (
    ClientBasicUpdateForm,
    ClientForm,
    ConsultorForm,
    ExitForm,
    ImportClientsForm,
    ImportMotivosRazoesForm,
    ImportResponsaveisForm,
    MotivoForm,
    RazaoForm,
    ResponsavelForm,
    ReuniaoPreferenciaForm,
    TransferForm,
    TermometroChangeForm,
    ValorChangeForm,
)
from .models import (
    AgendamentoAlinhamento,
    AgendamentoFechamento,
    Client,
    ClientHistory,
    Consultor,
    Motivo,
    Razao,
    Responsavel,
    ReuniaoPreferencia,
)
from .utils import build_operator_reports


def _normalize_text(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text


def _get_responsavel_suggestions() -> Iterable[str]:
    return Responsavel.objects.order_by("nome").values_list("nome", flat=True)


def is_admin(user):
    return user.is_superuser or user.groups.filter(name="Administrador").exists()


def _get_consultor_choices() -> Iterable[Consultor]:
    return Consultor.objects.order_by("nome")


def _get_motivo_suggestions() -> Iterable[str]:
    return Motivo.objects.order_by("nome").values_list("nome", flat=True)


def _get_razao_suggestions(tipo: str) -> Iterable[str]:
    return (
        Razao.objects.filter(tipo_de_historico=tipo)
        .order_by("nome")
        .values_list("nome", flat=True)
    )


def _ensure_motivo_entry(nome: str | None) -> Motivo | None:
    nome = (nome or "").strip()
    if not nome:
        return None
    motivo, _ = Motivo.objects.get_or_create(nome=nome)
    return motivo


def _ensure_razao_entry(nome: str | None, tipo: str, motivo: Motivo | None) -> None:
    nome = (nome or "").strip()
    if not nome:
        return
    motivo = motivo or _ensure_motivo_entry(nome)
    if motivo is None:
        return
    razao_obj, created = Razao.objects.get_or_create(
        nome=nome,
        tipo_de_historico=tipo,
        defaults={"motivo": motivo},
    )
    if not created and razao_obj.motivo_id != motivo.id:
        razao_obj.motivo = motivo
        razao_obj.save(update_fields=["motivo"])


def _register_motivo_razao(motivo_nome: str | None, razao_nome: str | None, tipo: str | None) -> None:
    motivo = _ensure_motivo_entry(motivo_nome)
    if razao_nome and tipo:
        _ensure_razao_entry(razao_nome, tipo, motivo)


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    if request.user.groups.filter(name="Agendamento").exists() and not (request.user.is_superuser or request.user.groups.filter(name="Administrador").exists()):
        return redirect("clientes:agendamentos")

    clients = Client.objects.prefetch_related("historico").all()
    stats = {
        "total": clients.count(),
        "ativos": clients.filter(status="ATIVO").count(),
        "inativos": clients.filter(status="INATIVO").count(),
        "receita_ativa": clients.filter(status="ATIVO").aggregate(total=Sum("valor"))["total"] or 0,
        "responsaveis": Responsavel.objects.count(),
    }
    reports = build_operator_reports(clients)
    all_months = reports.get("months", [])
    default_start = all_months[0] if all_months else ""
    default_end = all_months[-1] if all_months else ""

    requested_start = request.GET.get("mes_inicio") or default_start
    requested_end = request.GET.get("mes_fim") or default_end

    if requested_start and requested_end:
        start_dt = _parse_month_value(requested_start) or _parse_month_value(default_start)
        end_dt = _parse_month_value(requested_end) or _parse_month_value(default_end or requested_start)
        if start_dt and end_dt and start_dt > end_dt:
            start_dt, end_dt = end_dt, start_dt
        start_month = f"{start_dt.year}-{start_dt.month:02d}" if start_dt else ""
        end_month = f"{end_dt.year}-{end_dt.month:02d}" if end_dt else ""
        visible_months = _month_range(start_month, end_month) if start_month and end_month else all_months
    else:
        start_month = requested_start
        end_month = requested_end
        visible_months = all_months
    if not visible_months and all_months:
        visible_months = all_months

    quantity_report_filtered = _filter_series_report(reports.get("quantity_report"), visible_months)
    value_report_filtered = _filter_series_report(reports.get("value_report"), visible_months)

    combined_report = None
    if quantity_report_filtered and value_report_filtered and visible_months:
        q_map = {row["name"]: row["series"] for row in quantity_report_filtered.get("rows", [])}
        v_map = {row["name"]: row["series"] for row in value_report_filtered.get("rows", [])}

        def zero_series(months: list[str], zero_value):
            return [
                {"month": mes, "cumulative": zero_value, "entries": zero_value, "exits": zero_value}
                for mes in months
            ]

        combined_rows = []
        all_names = sorted(set(q_map.keys()) | set(v_map.keys()))
        for name in all_names:
            q_series = q_map.get(name) or zero_series(visible_months, 0)
            v_series = v_map.get(name) or zero_series(visible_months, Decimal("0"))
            months_data = []
            for idx, mes in enumerate(visible_months):
                q_value = q_series[idx] if idx < len(q_series) else {"month": mes, "cumulative": 0, "entries": 0, "exits": 0}
                v_value = v_series[idx] if idx < len(v_series) else {
                    "month": mes,
                    "cumulative": Decimal("0"),
                    "entries": Decimal("0"),
                    "exits": Decimal("0"),
                }
                months_data.append(
                    {
                        "month": mes,
                        "quantity": q_value,
                        "value": v_value,
                    }
                )
            combined_rows.append({"name": name, "months": months_data})

        q_totals = quantity_report_filtered.get("monthly_totals", [])
        v_totals = value_report_filtered.get("monthly_totals", [])
        q_zero_total = {"cumulative": 0, "entries": 0, "exits": 0}
        v_zero_total = {"cumulative": Decimal("0"), "entries": Decimal("0"), "exits": Decimal("0")}
        combined_totals = []
        for idx, mes in enumerate(visible_months):
            q_total = q_totals[idx] if idx < len(q_totals) else q_zero_total
            v_total = v_totals[idx] if idx < len(v_totals) else v_zero_total
            combined_totals.append({"month": mes, "quantity": q_total, "value": v_total})

        combined_report = {"rows": combined_rows, "monthly_totals": combined_totals}

    client_cashflow_report = reports.get("client_cashflow_report")
    cashflow_filtered = None
    if client_cashflow_report and visible_months:
        month_to_index = {m: i for i, m in enumerate(client_cashflow_report["months"])}
        indices = [month_to_index[m] for m in visible_months if m in month_to_index]
        if indices:
            filtered_rows = []
            for row in client_cashflow_report["rows"]:
                values = [row["values"][i] for i in indices]
                total = sum((Decimal(val) for val in values), Decimal("0"))
                if total > 0:
                    filtered_rows.append({"name": row["name"], "values": values, "total": total})

            summary = client_cashflow_report["summary"]
            def slice_series(series):
                return [series[i] for i in indices]

            def list_sum(values):
                total = Decimal("0")
                for val in values:
                    total += val
                return total

            cashflow_filtered = {
                "months": visible_months,
                "rows": filtered_rows,
                "summary": {
                    "total_value": {
                        "per_month": slice_series(summary["total_value"]),
                        "total": list_sum(slice_series(summary["total_value"])),
                    },
                    "active": {
                        "count": {
                            "per_month": slice_series(summary["active"]["count"]),
                            "total": sum(slice_series(summary["active"]["count"])),
                        },
                        "value": {
                            "per_month": slice_series(summary["active"]["value"]),
                            "total": list_sum(slice_series(summary["active"]["value"])),
                        },
                    },
                    "entries": {
                        "count": {
                            "per_month": slice_series(summary["entries"]["count"]),
                            "total": sum(slice_series(summary["entries"]["count"])),
                        },
                        "value": {
                            "per_month": slice_series(summary["entries"]["value"]),
                            "total": list_sum(slice_series(summary["entries"]["value"])),
                        },
                    },
                    "exits": {
                        "count": {
                            "per_month": slice_series(summary["exits"]["count"]),
                            "total": sum(slice_series(summary["exits"]["count"])),
                        },
                        "value": {
                            "per_month": slice_series(summary["exits"]["value"]),
                            "total": list_sum(slice_series(summary["exits"]["value"])),
                        },
                    },
                },
            }
    recentes = ClientHistory.objects.select_related("client").order_by("-data")[:10]
    operator_chart_data = None
    if quantity_report_filtered:
        colors = [
            "#2563eb",
            "#7c3aed",
            "#fb7185",
            "#0ea5e9",
            "#f97316",
            "#059669",
            "#a855f7",
        ]
        datasets = []
        for idx, row in enumerate(quantity_report_filtered["rows"]):
            datasets.append(
                {
                    "label": row["name"],
                    "data": [float(value["cumulative"]) for value in row["series"]],
                    "borderColor": colors[idx % len(colors)],
                    "backgroundColor": colors[idx % len(colors)],
                    "tension": 0.3,
                    "fill": False,
                }
            )
        operator_chart_data = json.dumps(
            {
                "type": "line",
                "data": {"labels": visible_months, "datasets": datasets},
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "interaction": {"mode": "index", "intersect": False},
                    "scales": {
                        "y": {
                            "beginAtZero": True,
                            "title": {"display": True, "text": "Quantidade"},
                        }
                    },
                    "plugins": {
                        "legend": {"position": "top"},
                        "tooltip": {"enabled": True},
                    },
                },
            }
        )

    operator_value_chart_data = None
    if value_report_filtered:
        datasets = []
        colors = [
            "#d946ef",
            "#22d3ee",
            "#f97316",
            "#14b8a6",
            "#6366f1",
            "#f43f5e",
        ]
        for idx, row in enumerate(value_report_filtered["rows"]):
            datasets.append(
                {
                    "label": row["name"],
                    "data": [float(value["cumulative"]) for value in row["series"]],
                    "borderColor": colors[idx % len(colors)],
                    "backgroundColor": colors[idx % len(colors)],
                    "tension": 0.3,
                    "fill": False,
                }
            )
        operator_value_chart_data = json.dumps(
            {
                "type": "line",
                "data": {"labels": visible_months, "datasets": datasets},
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "interaction": {"mode": "index", "intersect": False},
                    "scales": {
                        "y": {
                            "beginAtZero": True,
                            "title": {"display": True, "text": "Valores"},
                        }
                    },
                    "plugins": {
                        "legend": {"position": "top"},
                        "tooltip": {"enabled": True},
                    },
                },
            }
        )

    context = {
        "stats": stats,
        "reports": reports,
        "display_months": visible_months,
        "quantity_report": quantity_report_filtered,
        "value_report": value_report_filtered,
        "combined_report": combined_report,
        "client_cashflow_report": cashflow_filtered,
        "operator_chart_data": operator_chart_data,
        "operator_value_chart_data": operator_value_chart_data,
        "month_filter": {"start": start_month, "end": end_month},
        "recentes": recentes,
    }
    return render(request, "clientes/dashboard.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def financeiro_view(request: HttpRequest) -> HttpResponse:
    """
    Placeholder view for the Financeiro page, matching the new navbar links.
    """
    return render(request, "clientes/financeiro_placeholder.html")


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def usuarios_view(request: HttpRequest) -> HttpResponse:
    """
    Placeholder view for the Usuários page, matching the new navbar links.
    """
    return render(request, "clientes/usuarios_placeholder.html")


def _filter_clients_queryset(request: HttpRequest):
    clients = Client.objects.all()
    search_nome = request.GET.get("nome", "")
    search_responsavel = request.GET.get("responsavel", "")
    filter_status = request.GET.get("status", "")
    filter_termometro = request.GET.get("termometro", "")
    date_type = request.GET.get("data_tipo", "")
    date_start_raw = request.GET.get("data_inicio", "")
    date_end_raw = request.GET.get("data_fim", "")
    value_min_raw = request.GET.get("valor_min", "")
    value_max_raw = request.GET.get("valor_max", "")

    def _safe_parse_date(value: str) -> date | None:
        if not value:
            return None
        try:
            return _parse_date_value(value)
        except ValueError:
            return None

    def _safe_parse_decimal(value: str) -> Decimal | None:
        if not value:
            return None
        try:
            return _parse_decimal_value(value)
        except ValueError:
            return None

    date_start = _safe_parse_date(date_start_raw)
    date_end = _safe_parse_date(date_end_raw)
    value_min = _safe_parse_decimal(value_min_raw)
    value_max = _safe_parse_decimal(value_max_raw)

    if search_nome:
        clients = clients.filter(nome__icontains=search_nome)
    if search_responsavel:
        clients = clients.filter(responsavel__icontains=search_responsavel)
    if filter_status in {"ATIVO", "INATIVO"}:
        clients = clients.filter(status=filter_status)
    if filter_termometro and filter_termometro.isdigit():
        clients = clients.filter(termometro=int(filter_termometro))
    if date_type in {"entrada", "saida"}:
        date_field = "entrada" if date_type == "entrada" else "saida"
        if date_start:
            clients = clients.filter(**{f"{date_field}__gte": date_start})
        if date_end:
            clients = clients.filter(**{f"{date_field}__lte": date_end})
    if value_min is not None:
        clients = clients.filter(valor__gte=value_min)
    if value_max is not None:
        clients = clients.filter(valor__lte=value_max)

    filters = {
        "nome": search_nome,
        "responsavel": search_responsavel,
        "status": filter_status,
        "termometro": filter_termometro,
        "data_tipo": date_type,
        "data_inicio": date_start_raw,
        "data_fim": date_end_raw,
        "valor_min": value_min_raw,
        "valor_max": value_max_raw,
    }
    return clients.order_by("nome"), filters


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def client_list(request: HttpRequest) -> HttpResponse:
    clients, filters = _filter_clients_queryset(request)
    query_string = request.GET.urlencode()
    export_url = reverse("clientes:client_export")
    if query_string:
        export_url = f"{export_url}?{query_string}"
    context = {
        "clients": clients,
        "filters": filters,
        "responsaveis": _get_responsavel_suggestions(),
        "export_url": export_url,
    }
    return render(request, "clientes/client_list.html", context)


@login_required
def reunioes_lista(request: HttpRequest) -> HttpResponse:
    alinhamentos_qs = (
        ReuniaoPreferencia.objects.filter(tipo="ALINHAMENTO")
        .select_related("client")
        .order_by("client__nome")
    )
    fechamentos_qs = (
        ReuniaoPreferencia.objects.filter(tipo="FECHAMENTO")
        .select_related("client", "consultor")
        .order_by("client__nome")
    )
    active_clients = Client.objects.filter(status="ATIVO")

    alinhamentos = [
        {"client": pref.client, "pref": pref, "is_placeholder": False}
        for pref in alinhamentos_qs
    ]
    alinhamentos.extend(
        {
            "client": client,
            "pref": None,
            "is_placeholder": True,
            "message": "Cadastre Preferencias de Alinhamento",
        }
        for client in active_clients.filter(quer_alinhamento=True).exclude(
            pk__in=alinhamentos_qs.values_list("client_id", flat=True)
        )
    )
    alinhamentos.sort(key=lambda item: item["client"].nome.lower())

    fechamentos = [
        {"client": pref.client, "pref": pref, "is_placeholder": False}
        for pref in fechamentos_qs
    ]
    fechamentos.extend(
        {
            "client": client,
            "pref": None,
            "is_placeholder": True,
            "message": "Cadastre Preferencias de Fechamento",
        }
        for client in active_clients.exclude(pk__in=fechamentos_qs.values_list("client_id", flat=True))
    )
    fechamentos.sort(key=lambda item: item["client"].nome.lower())

    context = {
        "alinhamentos": alinhamentos,
        "fechamentos": fechamentos,
    }
    return render(request, "clientes/reunioes_lista.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def client_export(request: HttpRequest) -> HttpResponse:
    clients, _ = _filter_clients_queryset(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clientes"

    headers = [
        "Nome",
        "Responsável",
        "Status",
        "Termômetro",
        "Entrada",
        "Saída",
        "Valor",
        "Permuta",
        "Motivo",
        "Razão",
    ]
    worksheet.append(headers)

    for client in clients:
        worksheet.append(
            [
                client.nome,
                client.responsavel,
                client.status,
                client.termometro,
                client.entrada,
                client.saida,
                client.valor,
                "Sim" if client.permuta else "Não",
                client.motivo,
                client.razao,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="clientes_{timestamp}.xlsx"'
    workbook.save(response)
    return response


@login_required
def reunioes_export(request: HttpRequest) -> HttpResponse:
    alinhamentos_qs = (
        ReuniaoPreferencia.objects.filter(tipo="ALINHAMENTO")
        .select_related("client")
        .order_by("client__nome")
    )
    fechamentos_qs = (
        ReuniaoPreferencia.objects.filter(tipo="FECHAMENTO")
        .select_related("client", "consultor")
        .order_by("client__nome")
    )
    active_clients = Client.objects.filter(status="ATIVO")
 
    alinhamentos = [
        {"client": pref.client, "pref": pref, "is_placeholder": False}
        for pref in alinhamentos_qs
    ]
    alinhamentos.extend(
        {
            "client": client,
            "pref": None,
            "is_placeholder": True,
            "message": "Cadastre Preferencias de Alinhamento",
        }
        for client in active_clients.filter(quer_alinhamento=True).exclude(
            pk__in=alinhamentos_qs.values_list("client_id", flat=True)
        )
    )
    alinhamentos.sort(key=lambda item: item["client"].nome.lower())
 
    fechamentos = [
        {"client": pref.client, "pref": pref, "is_placeholder": False}
        for pref in fechamentos_qs
    ]
    fechamentos.extend(
        {
            "client": client,
            "pref": None,
            "is_placeholder": True,
            "message": "Cadastre Preferencias de Fechamento",
        }
        for client in active_clients.exclude(pk__in=fechamentos_qs.values_list("client_id", flat=True))
    )
    fechamentos.sort(key=lambda item: item["client"].nome.lower())
    
    workbook = Workbook()
    
    # Sheet Alinhamento
    ws_alinhamento = workbook.active
    ws_alinhamento.title = "Alinhamento"
    
    headers_alinhamento = [
        "Cliente", "Responsável", "Período", "Dia da semana", 
        "Horário", "Local", "Duração", "Dia sugerido", 
        "Observações", "Atualizado"
    ]
    ws_alinhamento.append(headers_alinhamento)
    
    for item in alinhamentos:
        client = item["client"]
        pref = item["pref"]
        
        if pref:
            periodo = f"{pref.dia_pref_inicio or ''} à {pref.dia_pref_fim or ''}" if pref.dia_pref_inicio and pref.dia_pref_fim else str(pref.dia_pref_inicio or "")
            if not pref.dia_pref_inicio and not pref.dia_pref_fim:
                periodo = "—"
                
            ws_alinhamento.append([
                client.nome,
                pref.responsavel_nome or client.responsavel or "—",
                periodo,
                pref.get_dia_semana_pref_display() or "—",
                pref.get_horario_pref_display() or "—",
                pref.get_local_display() or "—",
                f"{pref.duracao_minutos} min" if pref.duracao_minutos else "—",
                str(pref.data_sugerida) if pref.data_sugerida else "—",
                pref.observacoes or "—",
                pref.atualizado_em.strftime("%d/%m/%Y %H:%M") if pref.atualizado_em else "—"
            ])
        else:
            ws_alinhamento.append([
                client.nome,
                client.responsavel or "—",
                item["message"], # Período com mensagem de placeholder
                "—", "—", "—", "—", "—", "—", "—"
            ])
 
    # Sheet Fechamento
    ws_fechamento = workbook.create_sheet(title="Fechamento")
    
    headers_fechamento = [
        "Cliente", "Consultor", "Período", "Dia da semana", 
        "Horário", "Local", "Duração", "Dia sugerido", 
        "Observações", "Atualizado"
    ]
    ws_fechamento.append(headers_fechamento)
    
    for item in fechamentos:
        client = item["client"]
        pref = item["pref"]
        
        if pref:
            periodo = f"{pref.dia_pref_inicio or ''} à {pref.dia_pref_fim or ''}" if pref.dia_pref_inicio and pref.dia_pref_fim else str(pref.dia_pref_inicio or "")
            if not pref.dia_pref_inicio and not pref.dia_pref_fim:
                periodo = "—"
                
            consultor = str(pref.consultor) if pref.consultor else (pref.responsavel_nome or "—")
            
            ws_fechamento.append([
                client.nome,
                consultor,
                periodo,
                pref.get_dia_semana_pref_display() or "—",
                pref.get_horario_pref_display() or "—",
                pref.get_local_display() or "—",
                f"{pref.duracao_minutos} min" if pref.duracao_minutos else "—",
                str(pref.data_sugerida) if pref.data_sugerida else "—",
                pref.observacoes or "—",
                pref.atualizado_em.strftime("%d/%m/%Y %H:%M") if pref.atualizado_em else "—"
            ])
        else:
            ws_fechamento.append([
                client.nome,
                client.responsavel or "—",
                item["message"], # Período com mensagem de placeholder
                "—", "—", "—", "—", "—", "—", "—"
            ])
 
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = datetime.now().strftime("%Y%m%d")
    response["Content-Disposition"] = f'attachment; filename="reunioes_preferencias_{timestamp}.xlsx"'
    workbook.save(response)
    return response


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def client_create(request: HttpRequest) -> HttpResponse:
    responsaveis_choices = list(_get_responsavel_suggestions())
    if request.method == "POST":
        form = ClientForm(
            request.POST,
            responsavel_choices=responsaveis_choices,
            include_exit_fields=False,
        )
        if form.is_valid():
            client = form.save()
            messages.success(request, "Cliente cadastrado com sucesso.")
            return redirect("clientes:client_list")
    else:
        form = ClientForm(
            initial={"entrada": date.today()},
            responsavel_choices=responsaveis_choices,
            include_exit_fields=False,
        )
    context = {"form": form, "responsaveis": responsaveis_choices, "titulo": "Novo Cliente"}
    return render(request, "clientes/client_form.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def client_update(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        form = ClientBasicUpdateForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente atualizado.")
            return redirect("clientes:client_list")
    else:
        form = ClientBasicUpdateForm(instance=client)
    context = {"form": form, "responsaveis": [], "titulo": f"Editar {client.nome}"}
    return render(request, "clientes/client_form.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def client_delete(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        client.delete()
        messages.warning(request, "Cliente removido.")
        return redirect("clientes:client_list")
    return render(request, "clientes/confirm_delete.html", {"client": client})


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def transfer_client(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    responsaveis_choices = list(_get_responsavel_suggestions())
    motivos_transferencia = list(_get_motivo_suggestions())
    razoes_transferencia = list(_get_razao_suggestions("transferencia"))
    if request.method == "POST":
        form = TransferForm(
            request.POST,
            responsavel_choices=responsaveis_choices,
            motivo_choices=motivos_transferencia,
            razao_choices=razoes_transferencia,
        )
        if form.is_valid():
            novo_responsavel = form.cleaned_data["novo_responsavel"]
            motivo = form.cleaned_data["motivo"]
            razao = (form.cleaned_data.get("razao") or "").strip()
            data = form.cleaned_data["data"]
            ClientHistory.objects.create(
                client=client,
                tipo="TRANSFERENCIA",
                responsavel_antigo=client.responsavel,
                responsavel_novo=novo_responsavel,
                data=data,
                motivo=motivo,
                razao=razao,
            )
            client.responsavel = novo_responsavel
            client.save()
            Responsavel.objects.get_or_create(nome=novo_responsavel)
            ReuniaoPreferencia.objects.filter(client=client, tipo="ALINHAMENTO").update(
                responsavel_nome=novo_responsavel
            )
            _register_motivo_razao(motivo, razao, "transferencia")
            messages.success(request, "Transferência registrada com sucesso.")
            return redirect("clientes:client_list")
    else:
        form = TransferForm(
            initial={"novo_responsavel": client.responsavel},
            responsavel_choices=responsaveis_choices,
            motivo_choices=motivos_transferencia,
            razao_choices=razoes_transferencia,
        )

    context = {
        "form": form,
        "client": client,
        "responsaveis": responsaveis_choices,
        "motivos_transferencia": motivos_transferencia,
        "razoes_transferencia": razoes_transferencia,
    }
    return render(request, "clientes/transfer_form.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def exit_client(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    razoes_saida = list(_get_razao_suggestions("registro_de_saida"))
    if request.method == "POST":
        form = ExitForm(request.POST, razao_choices=razoes_saida)
        if form.is_valid():
            data = form.cleaned_data["data"]
            motivo = form.cleaned_data["motivo_saida"]
            razao = form.cleaned_data["razao_saida"]
            ClientHistory.objects.create(
                client=client,
                tipo="SAIDA",
                data=data,
                motivo=motivo,
                razao=razao,
                status_antigo=client.status,
                status_novo="INATIVO",
            )
            client.status = "INATIVO"
            client.saida = data
            client.motivo = motivo
            client.razao = razao
            client.save()
            _register_motivo_razao(motivo, razao, "registro_de_saida")
            messages.success(request, "Saída registrada.")
            return redirect("clientes:client_list")
    else:
        form = ExitForm(
            initial={"data": client.saida or date.today(), "motivo_saida": client.motivo, "razao_saida": client.razao},
            razao_choices=razoes_saida,
        )

    context = {
        "form": form,
        "client": client,
        "motivos_saida": _get_motivo_suggestions(),
        "razoes_saida": razoes_saida,
    }
    return render(request, "clientes/exit_form.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def change_termometro(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    razoes_termometro = list(_get_razao_suggestions("alteracao_de_termometro"))
    if request.method == "POST":
        form = TermometroChangeForm(request.POST, razao_choices=razoes_termometro)
        if form.is_valid():
            novo_termometro = int(form.cleaned_data["novo_termometro"])
            data = form.cleaned_data["data"]
            motivo = form.cleaned_data["motivo"]
            razao = form.cleaned_data["razao"]
            ClientHistory.objects.create(
                client=client,
                tipo="TERMOMETRO",
                data=data,
                motivo=motivo,
                razao=razao,
                termometro_antigo=client.termometro,
                termometro_novo=novo_termometro,
            )
            client.termometro = novo_termometro
            client.save(update_fields=["termometro"])
            _register_motivo_razao(motivo, razao, "alteracao_de_termometro")
            messages.success(request, "Alteração de termômetro registrada.")
            return redirect("clientes:client_list")
    else:
        form = TermometroChangeForm(
            initial={
                "novo_termometro": str(client.termometro),
                "data": date.today(),
            },
            razao_choices=razoes_termometro,
        )
    context = {
        "form": form,
        "client": client,
        "motivos_saida": _get_motivo_suggestions(),
        "razoes_saida": razoes_termometro,
    }
    return render(request, "clientes/termometro_form.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def change_valor(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    razoes_valor = list(_get_razao_suggestions("alteracao_de_valor"))
    if request.method == "POST":
        form = ValorChangeForm(request.POST, razao_choices=razoes_valor)
        if form.is_valid():
            valor = form.cleaned_data["valor"]
            permuta = form.cleaned_data["permuta"]
            data = form.cleaned_data["data"]
            motivo = form.cleaned_data["motivo"]
            razao = form.cleaned_data["razao"]
            ClientHistory.objects.create(
                client=client,
                tipo="VALOR",
                data=data,
                motivo=motivo,
                razao=razao,
                valor_antigo=client.valor,
                valor_novo=valor,
                permuta_antiga=client.permuta,
                permuta_nova=permuta,
            )
            client.valor = valor
            client.permuta = permuta
            client.save(update_fields=["valor", "permuta"])
            _register_motivo_razao(motivo, razao, "alteracao_de_valor")
            messages.success(request, "Alteração de valor registrada.")
            return redirect("clientes:client_list")
    else:
        form = ValorChangeForm(
            initial={
                "valor": client.valor,
                "permuta": client.permuta,
                "data": date.today(),
            },
            razao_choices=razoes_valor,
        )
    context = {
        "form": form,
        "client": client,
        "motivos_saida": _get_motivo_suggestions(),
        "razoes_saida": razoes_valor,
    }
    return render(request, "clientes/valor_form.html", context)


def _parse_date_value(value) -> date | None:
    if value in (None, "", "-", "N/A", "NA"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("\\", "/").replace("-", "/")
    if not text or text.upper() in {"NA", "N/A"}:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Data em formato inválido: {value}")


def _parse_decimal_value(value) -> Decimal:
    if value in (None, "", "N/A"):
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value)
    text = text.replace("R$", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    text = text.strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except Exception as exc:  # pragma: no cover
        raise ValueError(f"Valor inválido: {value}") from exc


def _parse_permuta(value) -> bool:
    text = _normalize_text(value).upper().replace(" ", "")
    if not text:
        return False
    return text in {"SIM", "TRUE", "1"}


def _parse_boolean_flag(value, default=True) -> bool:
    text = _normalize_text(value).upper().replace(" ", "")
    if not text:
        return default
    if text in {"SIM", "TRUE", "1", "ATIVO"}:
        return True
    if text in {"NAO", "NÃO", "FALSE", "0", "INATIVO"}:
        return False
    return default


def _parse_month_value(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m")
    except ValueError:
        return None


def _month_range(start: str, end: str) -> list[str]:
    start_dt = _parse_month_value(start)
    end_dt = _parse_month_value(end)
    if not start_dt or not end_dt or start_dt > end_dt:
        return []
    months = []
    current = datetime(start_dt.year, start_dt.month, 1)
    while current <= end_dt:
        months.append(f"{current.year}-{current.month:02d}")
        year = current.year + (1 if current.month == 12 else 0)
        month = 1 if current.month == 12 else current.month + 1
        current = datetime(year, month, 1)
    return months


def _filter_series_report(report: Dict[str, object] | None, months: list[str]):
    if not report or not months:
        return None

    def zero_like(example):
        return Decimal("0") if isinstance(example, Decimal) else 0

    filtered_rows = []
    for row in report.get("rows", []):
        series = row.get("series", [])
        if not series:
            continue
        zero_entry = zero_like(series[0]["entries"])
        zero_exit = zero_like(series[0]["exits"])
        zero_cumulative = zero_like(series[0]["cumulative"])

        last_cumulative = zero_cumulative
        for value in series:
            if value["month"] < months[0]:
                last_cumulative = value["cumulative"]
            else:
                break

        value_map = {value["month"]: value for value in series}
        filtered_series = []
        for mes in months:
            base_value = value_map.get(mes)
            if base_value:
                last_cumulative = base_value["cumulative"]
                entries = base_value["entries"]
                exits = base_value["exits"]
            else:
                entries = zero_entry
                exits = zero_exit
            filtered_series.append(
                {"month": mes, "cumulative": last_cumulative, "entries": entries, "exits": exits}
            )
        filtered_rows.append({"name": row.get("name", "—"), "series": filtered_series})

    totals = report.get("monthly_totals", [])
    zero_total = zero_like(totals[0]["cumulative"]) if totals else 0
    zero_total_entry = zero_like(totals[0]["entries"]) if totals else 0
    zero_total_exit = zero_like(totals[0]["exits"]) if totals else 0
    last_total_cumulative = zero_total
    for total in totals:
        if total["month"] < months[0]:
            last_total_cumulative = total["cumulative"]
        else:
            break
    totals_map = {total["month"]: total for total in totals}
    filtered_totals = []
    for mes in months:
        base_total = totals_map.get(mes)
        if base_total:
            last_total_cumulative = base_total["cumulative"]
            entries = base_total["entries"]
            exits = base_total["exits"]
        else:
            entries = zero_total_entry
            exits = zero_total_exit
        filtered_totals.append(
            {"month": mes, "cumulative": last_total_cumulative, "entries": entries, "exits": exits}
        )

    return {"rows": filtered_rows, "monthly_totals": filtered_totals}


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def import_clients(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ImportClientsForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = form.cleaned_data["arquivo"]
            try:
                arquivo.seek(0)
                wb = load_workbook(filename=arquivo, data_only=True)
                sheet = wb.active
            except Exception as exc:  # pragma: no cover
                return render(
                    request,
                    "clientes/import_form.html",
                    {
                        "form": ImportClientsForm(),
                        "import_errors": [f"Não foi possível ler o arquivo: {exc}"],
                    },
                )

            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                return render(
                    request,
                    "clientes/import_form.html",
                    {
                        "form": ImportClientsForm(),
                        "import_errors": ["Planilha vazia ou sem dados."],
                    },
                )

            header = rows[0]
            header_map: Dict[str, int] = {}
            for idx, value in enumerate(header):
                key = _normalize_text(value).upper().replace(" ", "")
                if key:
                    header_map[key] = idx

            required_columns = {"CLIENTE", "ENTRADA"}
            missing_columns = [col for col in required_columns if col not in header_map]
            if missing_columns:
                return render(
                    request,
                    "clientes/import_form.html",
                    {
                        "form": ImportClientsForm(),
                        "import_errors": [
                            f"A planilha deve conter as colunas: {', '.join(missing_columns)}."
                        ],
                    },
                )

            pending: List[Dict[str, object]] = []
            errors: List[str] = []

            for row_index, row in enumerate(rows[1:], start=2):
                if not row or all(cell in (None, "") for cell in row):
                    continue

                def get_value(key: str):
                    idx = header_map.get(key)
                    if idx is None or idx >= len(row):
                        return None
                    return row[idx]

                try:
                    nome = str(get_value("CLIENTE") or "").strip()
                    if not nome:
                        raise ValueError("Cliente não informado.")

                    responsavel = str(get_value("RESPONSAVEL") or "").strip() or "SEM RESPONSÁVEL"
                    termometro_raw = get_value("TERMOMETRO") or 3
                    try:
                        termometro = int(str(termometro_raw).strip())
                    except (ValueError, TypeError):
                        termometro = 3

                    status = str(get_value("STATUS") or "ATIVO").strip().upper()
                    if status not in {"ATIVO", "INATIVO"}:
                        status = "ATIVO"

                    entrada = _parse_date_value(get_value("ENTRADA"))
                    if not entrada:
                        raise ValueError("Data de entrada obrigatória.")

                    saida = _parse_date_value(get_value("SAIDA"))

                    permuta = _parse_permuta(get_value("PERMUTA"))
                    if permuta:
                        valor = Decimal("0")
                    else:
                        raw_valor = get_value("VALOR")
                        if status == "INATIVO" and (raw_valor in (None, "", "N/A")):
                            valor = Decimal("0")
                        else:
                            valor = _parse_decimal_value(raw_valor)
                            if status != "INATIVO" and valor <= 0:
                                raise ValueError("Valor deve ser maior que zero para clientes ativos sem permuta.")

                    motivo = str(get_value("MOTIVO") or "").strip()
                    razao = str(get_value("RAZAO") or "").strip()

                    pending.append(
                        {
                            "nome": nome,
                            "responsavel": responsavel,
                            "termometro": termometro,
                            "status": status,
                            "entrada": entrada,
                            "saida": saida,
                            "valor": valor,
                            "permuta": permuta,
                            "motivo": motivo,
                            "razao": razao,
                        }
                    )
                except Exception as exc:
                    errors.append(f"Linha {row_index}: {exc}")

            if errors:
                return render(
                    request,
                    "clientes/import_form.html",
                    {"form": ImportClientsForm(), "import_errors": errors},
                )

            with transaction.atomic():
                for data in pending:
                    client = Client.objects.create(**data)
                    Responsavel.objects.get_or_create(nome=client.responsavel)

            messages.success(request, f"{len(pending)} clientes importados.")
            return redirect("clientes:client_list")
    else:
        form = ImportClientsForm()
    return render(request, "clientes/import_form.html", {"form": form})


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def manage_responsaveis(request: HttpRequest) -> HttpResponse:
    import_form = ImportResponsaveisForm(prefix="import")
    if request.method == "POST":
        if "importar_responsaveis" in request.POST:
            import_form = ImportResponsaveisForm(request.POST, request.FILES, prefix="import")
            if import_form.is_valid():
                arquivo = import_form.cleaned_data["arquivo"]
                try:
                    arquivo.seek(0)
                    wb = load_workbook(filename=arquivo, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if len(rows) < 2:
                        raise ValueError("Planilha sem dados.")

                    header = rows[0]
                    header_map: Dict[str, int] = {}
                    for idx, value in enumerate(header):
                        key = _normalize_text(value).upper().replace(" ", "")
                        if key:
                            header_map[key] = idx

                    if "NOME" not in header_map:
                        raise ValueError("A coluna NOME é obrigatória.")

                    importados = 0
                    with transaction.atomic():
                        for row in rows[1:]:
                            if not row or all(cell in (None, "") for cell in row):
                                continue

                            def get_value(col: str):
                                index = header_map.get(col)
                                if index is None or index >= len(row):
                                    return None
                                return row[index]

                            nome = str(get_value("NOME") or "").strip()
                            if not nome:
                                continue
                            email = str(get_value("EMAIL") or "").strip()
                            ativo = _parse_boolean_flag(get_value("ATIVO"), default=True)

                            Responsavel.objects.update_or_create(
                                nome=nome,
                                defaults={"email": email, "ativo": ativo},
                            )
                            importados += 1

                    messages.success(request, f"{importados} responsáveis importados/atualizados.")
                except Exception as exc:
                    messages.error(request, f"Erro ao importar responsáveis: {exc}")
                return redirect("clientes:responsaveis")
        if "remover" in request.POST:
            responsavel = get_object_or_404(Responsavel, pk=request.POST.get("remover"))
            responsavel.delete()
            messages.warning(request, "Responsável removido.")
            return redirect("clientes:responsaveis")
        if "editar" in request.POST:
            responsavel = get_object_or_404(Responsavel, pk=request.POST.get("editar"))
            edit_form = ResponsavelForm(request.POST, instance=responsavel)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "Responsável atualizado.")
            else:
                messages.error(request, "Não foi possível atualizar o responsável.")
            return redirect("clientes:responsaveis")
        form = ResponsavelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Responsável cadastrado.")
            return redirect("clientes:responsaveis")
    else:
        form = ResponsavelForm()

    context = {
        "form": form,
        "import_form": import_form,
        "responsaveis": Responsavel.objects.order_by("nome"),
    }
    return render(request, "clientes/responsaveis.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def manage_consultores(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        if "remover" in request.POST:
            consultor = get_object_or_404(Consultor, pk=request.POST.get("remover"))
            consultor.delete()
            messages.warning(request, "Consultor removido.")
            return redirect("clientes:consultores")
        if "editar" in request.POST:
            consultor = get_object_or_404(Consultor, pk=request.POST.get("editar"))
            edit_form = ConsultorForm(request.POST, instance=consultor)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "Consultor atualizado.")
            else:
                messages.error(request, "Não foi possível atualizar o consultor.")
            return redirect("clientes:consultores")
        form = ConsultorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Consultor cadastrado.")
            return redirect("clientes:consultores")
    else:
        form = ConsultorForm()

    context = {
        "form": form,
        "consultores": _get_consultor_choices(),
    }
    return render(request, "clientes/consultores.html", context)


@login_required
def manage_reuniao_preferencias(request: HttpRequest, pk: int) -> HttpResponse:
    client = get_object_or_404(Client, pk=pk)
    pref_alinhamento = ReuniaoPreferencia.objects.filter(client=client, tipo="ALINHAMENTO").first()
    pref_fechamento = ReuniaoPreferencia.objects.filter(client=client, tipo="FECHAMENTO").first()

    form_type = request.POST.get("form_type") if request.method == "POST" else None
    form_alinhamento = ReuniaoPreferenciaForm(
        request.POST if form_type == "alinhamento" else None,
        prefix="alinhamento",
        instance=pref_alinhamento,
        client=client,
        tipo="ALINHAMENTO",
    )
    form_fechamento = ReuniaoPreferenciaForm(
        request.POST if form_type == "fechamento" else None,
        prefix="fechamento",
        instance=pref_fechamento,
        client=client,
        tipo="FECHAMENTO",
        require_alignment=client.quer_alinhamento and pref_alinhamento is None,
    )

    if request.method == "POST":
        if form_type == "alinhamento":
            if form_alinhamento.is_valid():
                pref = form_alinhamento.save(commit=False)
                pref.client = client
                pref.tipo = "ALINHAMENTO"
                pref.responsavel_nome = client.responsavel
                pref.consultor = None
                pref.save()
                messages.success(request, "Preferências de alinhamento salvas.")
                return redirect("clientes:reuniao_preferencias", pk=client.pk)
            else:
                messages.error(request, "Corrija os campos do alinhamento.")
        elif form_type == "fechamento":
            if form_fechamento.is_valid():
                pref = form_fechamento.save(commit=False)
                pref.client = client
                pref.tipo = "FECHAMENTO"
                if pref.consultor:
                    pref.responsavel_nome = pref.consultor.nome
                pref.save()
                messages.success(request, "Preferências de fechamento salvas.")
                return redirect("clientes:reuniao_preferencias", pk=client.pk)
            else:
                messages.error(request, "Corrija os campos do fechamento.")

    context = {
        "client": client,
        "form_alinhamento": form_alinhamento,
        "form_fechamento": form_fechamento,
        "pref_alinhamento": pref_alinhamento,
        "pref_fechamento": pref_fechamento,
    }
    return render(request, "clientes/reuniao_preferencias.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def manage_motivos_razoes(request: HttpRequest) -> HttpResponse:
    motivo_form = MotivoForm(prefix="motivo")
    razao_form = RazaoForm(prefix="razao")
    import_form = ImportMotivosRazoesForm(prefix="import")
    tipo_choices = dict(Razao.TIPO_HISTORICO_CHOICES)

    def _normalize_tipo(value: str | None) -> str | None:
        if not value:
            return None
        normalized = _normalize_text(value).lower().replace(" ", "_")
        mapping = {
            "transferencia": "transferencia",
            "transfer": "transferencia",
            "alteracao_de_valor": "alteracao_de_valor",
            "alteracaodevalor": "alteracao_de_valor",
            "valor": "alteracao_de_valor",
            "registro_de_saida": "registro_de_saida",
            "registrarsaida": "registro_de_saida",
            "saida": "registro_de_saida",
            "alteracao_de_termometro": "alteracao_de_termometro",
            "alteracaodetermometro": "alteracao_de_termometro",
            "termometro": "alteracao_de_termometro",
        }
        return mapping.get(normalized)

    if request.method == "POST":
        if "importar_motivos" in request.POST:
            import_form = ImportMotivosRazoesForm(request.POST, request.FILES, prefix="import")
            if import_form.is_valid():
                arquivo = import_form.cleaned_data["arquivo"]
                try:
                    arquivo.seek(0)
                    wb = load_workbook(filename=arquivo, data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    if len(rows) < 2:
                        raise ValueError("Planilha sem dados.")

                    header = rows[0]
                    header_map: Dict[str, int] = {}
                    for idx, value in enumerate(header):
                        key = _normalize_text(value).upper().replace(" ", "")
                        if key:
                            header_map[key] = idx

                    if not {"MOTIVO", "RAZAO", "TIPO", "MOTIVOTRANSFERENCIA"} & set(header_map.keys()):
                        raise ValueError("A planilha deve conter pelo menos uma das colunas MOTIVO ou RAZAO/TIPO.")

                    novos_motivos = set()
                    novas_razoes = 0

                    with transaction.atomic():
                        for row in rows[1:]:
                            if not row or all(cell in (None, "") for cell in row):
                                continue

                            def get_value(col: str):
                                index = header_map.get(col)
                                if index is None or index >= len(row):
                                    return None
                                return row[index]

                            motivo_nome = str(get_value("MOTIVO") or "").strip()
                            if motivo_nome:
                                novos_motivos.add(motivo_nome)
                                _ensure_motivo_entry(motivo_nome)

                            transf_nome = str(get_value("MOTIVOTRANSFERENCIA") or "").strip()
                            if transf_nome:
                                novos_motivos.add(transf_nome)
                                _ensure_motivo_entry(transf_nome)

                            razao_nome = str(get_value("RAZAO") or "").strip()
                            tipo_val = str(get_value("TIPO") or "").strip()
                            tipo_normalizado = _normalize_tipo(tipo_val)
                            if razao_nome and tipo_normalizado:
                                motivo_obj = _ensure_motivo_entry(motivo_nome)
                                _ensure_razao_entry(razao_nome, tipo_normalizado, motivo_obj)
                                novas_razoes += 1

                    messages.success(
                        request,
                        f"Importação concluída: {len(novos_motivos)} motivos e {novas_razoes} razões.",
                    )
                except Exception as exc:
                    messages.error(request, f"Erro na importação: {exc}")
                return redirect("clientes:motivos_razoes")

        if "remover_motivo" in request.POST:
            Motivo.objects.filter(pk=request.POST["remover_motivo"]).delete()
            messages.warning(request, "Motivo removido.")
            return redirect("clientes:motivos_razoes")
        if "editar_motivo" in request.POST:
            motivo = get_object_or_404(Motivo, pk=request.POST["editar_motivo"])
            nome = request.POST.get("nome_motivo", "").strip()
            if nome:
                motivo.nome = nome
                motivo.save(update_fields=["nome"])
                messages.success(request, "Motivo atualizado.")
            else:
                messages.error(request, "O nome do motivo é obrigatório.")
            return redirect("clientes:motivos_razoes")
        if "remover_razao" in request.POST:
            Razao.objects.filter(pk=request.POST["remover_razao"]).delete()
            messages.warning(request, "Razão removida.")
            return redirect("clientes:motivos_razoes")
        if "editar_razao" in request.POST:
            razao = get_object_or_404(Razao, pk=request.POST["editar_razao"])
            nome = request.POST.get("nome_razao", "").strip()
            tipo_val = request.POST.get("tipo_razao")
            motivo_id = request.POST.get("motivo_razao")
            if not nome:
                messages.error(request, "O nome da razão é obrigatório.")
                return redirect("clientes:motivos_razoes")
            if tipo_val not in tipo_choices:
                messages.error(request, "Tipo inválido para a razão.")
                return redirect("clientes:motivos_razoes")
            motivo = get_object_or_404(Motivo, pk=motivo_id) if motivo_id else razao.motivo
            razao.nome = nome
            razao.tipo_de_historico = tipo_val
            razao.motivo = motivo
            razao.save(update_fields=["nome", "tipo_de_historico", "motivo"])
            messages.success(request, "Razão atualizada.")
            return redirect("clientes:motivos_razoes")

        motivo_form = MotivoForm(request.POST, prefix="motivo")
        razao_form = RazaoForm(request.POST, prefix="razao")

        if motivo_form.is_valid():
            motivo_form.save()
            messages.success(request, "Motivo adicionado.")
            return redirect("clientes:motivos_razoes")
        if razao_form.is_valid():
            razao_form.save()
            messages.success(request, "Razão adicionada.")
            return redirect("clientes:motivos_razoes")

    context = {
        "motivo_form": motivo_form,
        "razao_form": razao_form,
        "import_form": import_form,
        "motivos": Motivo.objects.order_by("nome"),
        "razoes": Razao.objects.select_related("motivo").order_by("nome"),
        "tipo_choices": Razao.TIPO_HISTORICO_CHOICES,
    }
    return render(request, "clientes/motivos_razoes.html", context)


@user_passes_test(is_admin, login_url='clientes:acesso_negado')
def clear_database(request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        return redirect("clientes:dashboard")

    with transaction.atomic():
        ClientHistory.objects.all().delete()
        Client.objects.all().delete()
        Responsavel.objects.all().delete()
        Motivo.objects.all().delete()
        Razao.objects.all().delete()
    messages.warning(request, "Todos os registros foram removidos.")
    return redirect("clientes:dashboard")


@login_required
def agendamentos_view(request: HttpRequest) -> HttpResponse:
    return render(request, "clientes/agendamentos.html")


@login_required
def agendamentos_api_list(request: HttpRequest) -> JsonResponse:
    try:
        mes = int(request.GET.get("mes", date.today().month))
        ano = int(request.GET.get("ano", date.today().year))
    except ValueError:
        return JsonResponse({"error": "Mês/Ano inválidos"}, status=400)

    clients = Client.objects.filter(status="ATIVO").order_by("nome")
    
    # Prefetch preferences to avoid N+1
    clients = clients.prefetch_related("preferencias_reuniao")

    # Fetch existing agendamentos
    alinhamentos = {
        a.client_id: a 
        for a in AgendamentoAlinhamento.objects.filter(mes=mes, ano=ano)
    }
    fechamentos = {
        f.client_id: f 
        for f in AgendamentoFechamento.objects.filter(mes=mes, ano=ano)
    }

    data = []
    for client in clients:
        pref_alinhamento = client.preferencias_reuniao.filter(tipo="ALINHAMENTO").first()
        pref_fechamento = client.preferencias_reuniao.filter(tipo="FECHAMENTO").first()
        
        c_alinhamento = alinhamentos.get(client.id)
        c_fechamento = fechamentos.get(client.id)
        
        # Serialize preferences
        def serialize_pref(pref):
            if not pref: return None
            return {
                "dia_semana": pref.get_dia_semana_pref_display(),
                "horario": pref.get_horario_pref_display(),
                "local": pref.get_local_display(),
                "observacoes": pref.observacoes,
                "responsavel": pref.responsavel_nome or client.responsavel,
                "consultor": str(pref.consultor) if pref.consultor else None,
                "duracao": f"{pref.duracao_minutos} min" if pref.duracao_minutos else None,
                "dia_sugerido": str(pref.data_sugerida) if pref.data_sugerida else None,
            }

        item = {
            "client": {
                "id": client.id,
                "nome": client.nome,
                "quer_alinhamento": client.quer_alinhamento,
            },
            "prefs": {
                "alinhamento": serialize_pref(pref_alinhamento),
                "fechamento": serialize_pref(pref_fechamento),
            },
            "agendamento": {
                "alinhamento": {
                    "data": c_alinhamento.data_reuniao if c_alinhamento else None,
                    "horario": c_alinhamento.horario if c_alinhamento else "",
                    "status": c_alinhamento.status if c_alinhamento else "PENDENTE",
                    "observacao": c_alinhamento.observacao if c_alinhamento else "",
                },
                "fechamento": {
                    "data": c_fechamento.data_reuniao if c_fechamento else None,
                    "horario": c_fechamento.horario if c_fechamento else "",
                    "status": c_fechamento.status if c_fechamento else "PENDENTE",
                    "observacao": c_fechamento.observacao if c_fechamento else "",
                }
            }
        }
        data.append(item)

    return JsonResponse({"data": data})


@login_required
def agendamentos_api_save(request: HttpRequest) -> JsonResponse:
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    try:
        data = json.loads(request.body)
        tipo = data.get("tipo") # "alinhamento" or "fechamento"
        client_id = data.get("client_id")
        mes = int(data.get("mes"))
        ano = int(data.get("ano"))
        
        fields = {
            "data_reuniao": data.get("data") or None,
            "horario": data.get("horario", ""),
            "status": data.get("status", "PENDENTE"),
            "observacao": data.get("observacao", ""),
        }
        
        model = AgendamentoAlinhamento if tipo == "alinhamento" else AgendamentoFechamento
        
        obj, created = model.objects.update_or_create(
            client_id=client_id,
            mes=mes,
            ano=ano,
            defaults=fields
        )
        
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)
